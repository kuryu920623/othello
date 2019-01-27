from PC_Choice import *
from CulcBoardStateNumbers import *
from Othello_OneMatch import *

import random
import openpyxl as px
from copy import copy

PCNum = 13 #一世代に存在するPCの数
WB_Result = px.Workbook()
WB_Result.save("AllResult.xlsx")
WS_AllResult = WB_Result['Sheet']
#PC名記入、結果まとめシートの幅調整
for i in range(PCNum):
    WS_AllResult.cell(1,i+2,"PC" + str(i).zfill(2))
    WS_AllResult.column_dimensions[chr(ord("B")+i)].width = 7

#盤面のポイントに対して個さの違う色を返す
def ReturnPatternFill(point):
    if point < 0: #マイナス、青
        OtherVar = hex(200 + point)[-2:] #B以外の値は200~100
        return px.styles.PatternFill('solid',OtherVar + OtherVar +'FF', OtherVar + OtherVar +'FF')
    else: #プラスまたは0、赤
        OtherVar = hex(int(200 - point * 1.5))[-2:] #R以外の値は200~50
        return px.styles.PatternFill('solid','FF' + OtherVar + OtherVar,'FF' + OtherVar + OtherVar)

#ランダムな点数のPCを生成し、Excelにポイントに記録
def MakePC_and_WriteToExcel(PC_i,Sheet):
    Sheet.cell(PC_i*10+1,1,"PC" + str(PC_i).zfill(2))

    #tについて
    t1 = random.randint(-500,500)
    t2 = random.randint(-500,500)
    Sheet.cell(PC_i*10+2,1,t1)
    Sheet.cell(PC_i*10+3,1,t2)

    #マスごとのポイント
    point_array1 = [[0 for i in range(8)] for j in range(8)]
    point_array2 = [[0 for i in range(8)] for j in range(8)]
    for i in range(4):
        for j in range(i,4):
            point1_ij, point2_ij = random.randint(-100,100), random.randint(-100,100)
            for i2 in (i,7-i):
                for j2 in (j,7-j):
                    point_array1[i2][j2] = point_array1[j2][i2] = point1_ij
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2, point1_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2, point1_ij)
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2).fill = ReturnPatternFill(point1_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2).fill = ReturnPatternFill(point1_ij)

                    point_array2[i2][j2] = point_array2[j2][i2] = point2_ij
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2 +9, point2_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2 +9, point2_ij)
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2 +9).fill = ReturnPatternFill(point2_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2 +9).fill = ReturnPatternFill(point2_ij)

    PC = PC_Choice(t1,t2,point_array1,point_array2)
    return PC

for generation in range(1,20):
    if generation == 1: #第一世代だけ、勝者がいないため追加で別処理
        PCList_gene = [] #各PCインスタンスが格納されるリスト
        WS_Point_Gene = WB_Result.create_sheet("Gene" + str(generation).zfill(2) + "Point", generation) #1世代のポイントのシート作成
        PCList_gene.append(MakePC_and_WriteToExcel(0,WS_Point_Gene)) # インデックス0のPCを作成&リストに追加
        PCList_gene.append(MakePC_and_WriteToExcel(1,WS_Point_Gene)) # インデックス1のPCを作成&リストに追加
        PCList_gene.append(MakePC_and_WriteToExcel(2,WS_Point_Gene)) # インデックス2のPCを作成&リストに追加
    for PC_i in range(3,PCNum): #2~PCNum番までのPCを作成。1世代の時の0,1はこの前に作っている。
        PCList_gene.append(MakePC_and_WriteToExcel(PC_i,WS_Point_Gene))

    #見やすいように世代のポイント列幅の変更
    for i in range(ord('B'),ord('R')+1):
        WS_Point_Gene.column_dimensions[chr(i)].width = 4
    WB_Result.save("AllResult.xlsx")

    #対戦開始
    WS_Result_Gene = WB_Result.create_sheet("Gene" + str(generation).zfill(2) + "Result")
    WinPoint = [0 for i in range(PCNum)] #PC毎の勝ち点、勝利:+1,負け:-1,引き分け:±0
    MatchNum = 0
    for PC1 in range(PCNum):
        for PC2 in [i for i in range(PCNum) if i != PC1]:
            MatchNum += 1
            WS_Result_Gene.cell(MatchNum,1,PC1)
            WS_Result_Gene.cell(MatchNum,2,PC2)
            result_OneMatch = OneMatch(1,1,PCList_gene[PC1],PCList_gene[PC2]) #[黒の石数,白の石数,時間(s)]

            if result_OneMatch[0] > result_OneMatch[1]: #黒が勝ち
                WinPoint[PC1] += 1
                WinPoint[PC2] -= 1
                WS_Result_Gene.cell(MatchNum,3,PC1)
                WS_Result_Gene.cell(MatchNum,1).fill = ReturnPatternFill(50)
            elif result_OneMatch[1] > result_OneMatch[0]: #白が勝ち
                WinPoint[PC1] -= 1
                WinPoint[PC2] += 1
                WS_Result_Gene.cell(MatchNum,3,PC2) #色を変えたい
                WS_Result_Gene.cell(MatchNum,2).fill = ReturnPatternFill(50)
            #引き分けの場合両方0点

            WS_Result_Gene.cell(MatchNum,4,result_OneMatch[0])
            WS_Result_Gene.cell(MatchNum,5,result_OneMatch[1])
            WS_Result_Gene.cell(MatchNum,6,result_OneMatch[2])

            print(result_OneMatch)

    #世代の結果のまとめを記入
    WS_AllResult.cell(generation+1,1,generation)
    for i in range(PCNum):
        WS_AllResult.cell(generation+1,i+2,WinPoint[i])
    WB_Result.save("AllResult.xlsx")

    print(WinPoint)

    WinPoint_dic = [(i,WinPoint[i]) for i in range(PCNum)]

    FirstPC = sorted(WinPoint_dic,key = lambda x:x[1])[-1][0]
    SecondPC = sorted(WinPoint_dic,key = lambda x:x[1])[-2][0]
    WS_AllResult.cell(generation+1,FirstPC+2).fill = ReturnPatternFill(75)
    WS_AllResult.cell(generation+1,SecondPC+2).fill = ReturnPatternFill(25)


    PCList_gene = [PCList_gene[FirstPC],PCList_gene[SecondPC]] #勝者をPC0,1にする
    #勝者以外のインスタンスを削除
    newWS_Point_Gene = WB_Result.create_sheet("Gene" + str(generation+1).zfill(2) + "Point", generation+1) #1世代のポイントのシート作成
    #勝者のポイントを次世代のシートにコピー
    newWS_Point_Gene.cell(1,1).value = "PC0"
    newWS_Point_Gene.cell(11,1).value = "PC1"
    for i in range(8):
        for j in range(18):
            newWS_Point_Gene.cell(2+i,1+j).value =  WS_Point_Gene.cell(2+FirstPC*10 +i,1+j).value #一位のやつ
            newWS_Point_Gene.cell(2+i,1+j)._style = copy(WS_Point_Gene.cell(2+FirstPC*10 +i,1+j)._style)
            newWS_Point_Gene.cell(12+i,1+j).value = WS_Point_Gene.cell(2+SecondPC*10+i,1+j).value #二位のやつ
            newWS_Point_Gene.cell(12+i,1+j)._style = copy(WS_Point_Gene.cell(2+SecondPC*10+i,1+j)._style)
    WS_Point_Gene = newWS_Point_Gene

    #一位と二位のポイントの中間をとったPCを次の世代の3体目にする。
    #コードが汚いので後で変えたい
    Sheet = WS_Point_Gene
    Sheet.cell(PC_i*10+1,1,"PC2")
    PC_i = 2
    t1 = int(PCList_gene[0].t1/2 + PCList_gene[1].t1/2)
    t2 = int(PCList_gene[0].t2/2 + PCList_gene[1].t2/2)
    Sheet.cell(2*10+2,1,t1)
    Sheet.cell(2*10+3,1,t2)
    #マスごとのポイント
    point_array1 = [[0 for i in range(8)] for j in range(8)]
    point_array2 = [[0 for i in range(8)] for j in range(8)]
    for i in range(4):
        for j in range(i,4):
            point1_ij = int(WS_Point_Gene.cell(2+i,2+j).value/2 + WS_Point_Gene.cell(12+i,2+j).value/2)
            point2_ij = int(WS_Point_Gene.cell(2+i,11+j).value/2 + WS_Point_Gene.cell(12+i,11+j).value/2)
            for i2 in (i,7-i):
                for j2 in (j,7-j):
                    point_array1[i2][j2] = point_array1[j2][i2] = point1_ij
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2, point1_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2, point1_ij)
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2).fill = ReturnPatternFill(point1_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2).fill = ReturnPatternFill(point1_ij)

                    point_array2[i2][j2] = point_array2[j2][i2] = point2_ij
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2 +9, point2_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2 +9, point2_ij)
                    Sheet.cell(PC_i*10+2 +i2, 2 +j2 +9).fill = ReturnPatternFill(point2_ij)
                    Sheet.cell(PC_i*10+2 +j2, 2 +i2 +9).fill = ReturnPatternFill(point2_ij)

    PC = PC_Choice(t1,t2,point_array1,point_array2)
    PCList_gene.append(PC)

    #世代のインスタンスを、すべて削除
